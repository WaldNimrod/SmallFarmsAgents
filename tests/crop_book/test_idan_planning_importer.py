"""Tests for idan_planning_importer (L03/L04) — WP-C1 AC-C1-07."""
from pathlib import Path

import pytest

pytestmark = pytest.mark.crop_book

L03 = Path("data/external_sources/israeli/L03_IDAN_winter_planning.xlsx")
L04 = Path("data/external_sources/israeli/L04_IDAN_summer_planning.xlsx")


class TestIdanPlanningImporter:
    def test_l03_winter_sheet_rows(self):
        import openpyxl

        wb = openpyxl.load_workbook(L03, read_only=True, data_only=True)
        assert wb["תוכנית גידול"].max_row == 203
        wb.close()

    def test_l04_summer_sheet_rows(self):
        import openpyxl

        wb = openpyxl.load_workbook(L04, read_only=True, data_only=True)
        assert wb["תוכנית גידול"].max_row == 150
        wb.close()

    def test_skips_summary_rows(self):
        from organic_market_agent.crop_book.importer.israeli.idan_planning_importer import (
            parse_idan_xlsx,
        )

        sv, _ = parse_idan_xlsx(L03)
        crops = {r["crop_name_he"] for r in sv}
        assert "סיכום" not in crops

    def test_planting_date_parsing(self):
        from organic_market_agent.crop_book.importer.israeli.idan_planning_importer import (
            _parse_month_token,
        )

        assert _parse_month_token("13.4.2017") == 4
        assert _parse_month_token("נוב") == 11
        assert _parse_month_token("סוף אוקטובר (25)") == 10
