"""Script to generate minimal_masterclass.xlsx fixture for tests.

3 crops: Arugula, Carrots, Basil — covering all 5 sheets.
Run once to regenerate: python3 make_fixture.py
"""
from pathlib import Path
import openpyxl


def make_fixture():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # ---- Sheet 1: CROP CHART ----
    ws_cc = wb.create_sheet("CROP CHART")
    ws_cc.append(["CROP CHART", None, None, None, None, None])
    ws_cc.append([None] * 6)
    ws_cc.append([None] * 6)
    ws_cc.append([None] * 6)
    ws_cc.append(["Crop", "Rows per Bed", "Spacing", None, "DTM", "Harvest window",
                  None, None, "Yields / 100' bed", "Price/unit", "Unit", "Revenue"])
    ws_cc.append(["Arugula", 12, 2, 5, 30, 17, 47, 0, 105, 9, "Lbs", 945])
    ws_cc.append(["Carrots", 3, 5, 13, 50, 21, 46, 25, 160, 3, "Bunch", 480])
    ws_cc.append(["Basil", 4, 12, 30, 70, 14, 59, 25, 400, 3, "Head", 1200])

    # ---- Sheet 2: CROP ASSOCIATED TASKS ----
    ws_cat = wb.create_sheet("CROP ASSOCIATED TASKS")
    ws_cat.append([None, None, None, "Task, number of days after seeding"])
    ws_cat.append(["CODE", "CROP", "At seeding/transplanting", "Stale seed bed",
                    "flame weeder", "flextine harrow-1", "flextine harrow-2",
                    "Biodisc", "Hoe", "handweed", "Boron/seaweed -1", "Boron/seaweed -2",
                    "Straw mulch+topdress", "head pinch/chop", "mow and tarp"])
    ws_cat.append(["AR", "Arugula", "X", None, None, None, None, None, None, None, None, None, None, None, None])
    ws_cat.append(["CA", "Carrots", None, None, -7, None, None, None, None, 21, None, None, None, None, None])
    ws_cat.append(["BA", "Basil", None, None, None, None, None, None, None, None, None, None, None, 21, 21])

    # ---- Sheet 3: DIRECT SEEDING CHART ----
    ws_ds = wb.create_sheet("DIRECT SEEDING CHART")
    ws_ds.append(["CROP", "CULTIVAR", "ROWS PER BED", "SPACING / ROW", "SEEDER",
                   "CALIBRATION", "DENSITY TARGET /100 bed", "NOTES"])
    ws_ds.append(["Arugula", "Astro", 12, '1.5"', "6-row", "", "90 g", ""])
    ws_ds.append(["Carrots", "Napoli", 3, '2.5"', "Jang", "", "75 g", ""])
    # Basil is transplant — not in direct seeding chart

    # ---- Sheet 4: NURSERY & TRANSPLANT CHART ----
    ws_ns = wb.create_sheet("NURSERY & TRANSPLANT CHART")
    ws_ns.append(["CROP", "# Row", "Spacing on the row Inch", None, None,
                   None, "Tray number", "Days in cell", "Nursery Notes", "Transplant Notes"])
    ws_ns.append(["Basil", 4, 12, 30, None, None, "128", 30, None, "Grow in caterpillar"])
    ws_ns.append(["Carrots", 3, 5, 13, None, None, "128", 30, "Use new seeds only", None])

    # ---- Sheet 5: CULTIVARS ----
    ws_cv = wb.create_sheet("CULTIVARS")
    ws_cv.append(["CULTIVARS"])
    ws_cv.append([None])
    ws_cv.append([None])
    ws_cv.append(["Species", "Cultivar", "Provider", "Photo", "DTM",
                   "Information from the provider", "Comments"])
    ws_cv.append(["Carrots", "Napoli", "JSS", None, 58, "Ideal for overwintering.", None])
    ws_cv.append(["Carrots", "Bolero", "JSS", None, None, "Ideal storage carrot.", None])
    ws_cv.append(["Basil", "Genovese", "High Mowing", None, 70, "Standard Genovese.", "Classic basil"])
    ws_cv.append(["Arugula", "Astro", "JSS", None, 28, "Fast growing arugula.", None])

    out_path = Path(__file__).parent / "minimal_masterclass.xlsx"
    wb.save(str(out_path))
    print(f"Written: {out_path}")


if __name__ == "__main__":
    make_fixture()
