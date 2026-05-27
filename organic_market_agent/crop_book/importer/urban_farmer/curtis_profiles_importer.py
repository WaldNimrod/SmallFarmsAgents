"""Curtis Stone master chart importer (L40 XLSX) — WP-C3 §5.

Source: L40_curtis_crop_profiles.xlsx sheet 'Sheet 1 - Master Chart-1'
Target: crop_variety_source_values with source='OP:CurtisStone'
"""

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from organic_market_agent.crop_book.importer.urban_farmer._shared import (
    CURTIS_CROP_MAP,
    URBAN_FARMER_DIR,
    ImportSummary,
)

logger = logging.getLogger(__name__)

SOURCE = "OP:CurtisStone"
TRUST = "OP"
CONFIDENCE = Decimal("0.55")
_SHEET = "Sheet 1 - Master Chart-1"

# Column indices (0-based)
_COL_CROP = 0
_COL_DTM = 1
_COL_CVR = 4     # CVR5/5 rating
_COL_DS_TR = 6   # DS/TR (direct seed / transplant)
_COL_JANG = 13   # Jang Roller
_COL_EW = 14     # EW Plate
_COL_ROWS_BED = 15  # Rows/Bed
_COL_YIELD = 25  # Avg Yield/25' bed total

_DS_VALUES = {"ds", "direct seed", "d.s.", "ds/tr"}
_TR_VALUES = {"tr", "transplant", "trn", "t/r"}


def _parse_planting_method(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in _DS_VALUES:
        return "direct_sow"
    if s in _TR_VALUES:
        return "transplant"
    if "ds" in s and "tr" not in s:
        return "direct_sow"
    if "tr" in s or "trn" in s:
        return "transplant"
    return None


def _parse_int(raw: object) -> int | None:
    if raw is None:
        return None
    try:
        return int(raw)
    except (ValueError, TypeError):
        return None


def _parse_decimal(raw: object) -> Decimal | None:
    if raw is None:
        return None
    try:
        return Decimal(str(raw)).quantize(Decimal("0.001"))
    except InvalidOperation:
        return None


def _resolve_variety_id(session: Session, crop_name_en: str) -> int | None:
    """Return default variety_id for the crop, or None if crop not in DB."""
    from organic_market_agent.crop_book.models import Crop, CropVariety

    canonical_he = CURTIS_CROP_MAP.get(crop_name_en)
    if canonical_he is None:
        logger.debug("Curtis: no mapping for %r — skipping", crop_name_en)
        return None
    crop = session.query(Crop).filter_by(name_he=canonical_he).one_or_none()
    if crop is None:
        logger.debug("Curtis: crop %r not in DB — skipping", canonical_he)
        return None
    v = (
        session.query(CropVariety)
        .filter(CropVariety.crop_id == crop.id, CropVariety.name_en.is_(None))
        .order_by(CropVariety.id)
        .first()
    )
    if v is None:
        v = CropVariety(crop_id=crop.id, name_en=None, name_he=None)
        session.add(v)
        session.flush()
    return v.id


def _upsert(session: Session, variety_id: int, field_name: str,
            value_numeric: Decimal | None, value_text: str | None = None,
            unit: str | None = None, note: str | None = None,
            confidence_weight: Decimal | None = None) -> None:
    from organic_market_agent.crop_book.models import CropVarietySourceValue

    row = (
        session.query(CropVarietySourceValue)
        .filter_by(variety_id=variety_id, field_name=field_name, source=SOURCE)
        .one_or_none()
    )
    if row is None:
        row = CropVarietySourceValue(variety_id=variety_id, field_name=field_name, source=SOURCE)
        session.add(row)
    row.value_numeric = value_numeric
    row.value_text = value_text or (str(value_numeric) if value_numeric is not None else None)
    row.unit = unit
    row.trust_tier = TRUST
    row.confidence_weight = confidence_weight if confidence_weight is not None else CONFIDENCE
    row.is_outlier_rejected = False
    row.note = note
    session.flush()


def import_all(session: Session, xlsx_path: Path | None = None) -> ImportSummary:
    xlsx = xlsx_path or (URBAN_FARMER_DIR / "L40_curtis_crop_profiles.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[_SHEET]

    summary = ImportSummary()
    map_misses: list[str] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or row[_COL_CROP] is None:
            continue
        crop_name = str(row[_COL_CROP]).strip()
        if not crop_name or crop_name.upper() == "MICROGREENS":
            continue

        summary.rows_parsed += 1

        # Skip shoot crops (not plantable field crops)
        if any(kw in crop_name.lower() for kw in ("shoot", "flat")):
            continue

        variety_id = _resolve_variety_id(session, crop_name)
        if variety_id is None:
            map_misses.append(crop_name)
            continue

        upserted = False

        # DTM — confidence_weight=0 so North American context does not shift
        # the Israeli shadow calibration blend (F-C3-LV-01 remediation).
        # Data is preserved for reference; weighted_mean blend contribution = 0.
        dtm = _parse_int(row[_COL_DTM])
        if dtm is not None:
            cvr = row[_COL_CVR]
            note_str = f"curtis_rating:{cvr}" if cvr not in (None, "-", "") else None
            _upsert(session, variety_id, "days_to_maturity", Decimal(dtm), unit="days",
                    note=note_str, confidence_weight=Decimal("0"))
            upserted = True

        # Planting method
        pm = _parse_planting_method(row[_COL_DS_TR])
        if pm:
            _upsert(session, variety_id, "planting_method", None, value_text=pm)
            upserted = True

        # Seeder info (Jang roller + EW plate)
        jang = row[_COL_JANG]
        ew = row[_COL_EW]
        if jang not in (None, "-", ""):
            plate = ew if ew not in (None, "-", "") else "none"
            seeder_text = f"roller:{jang}:plate:{plate}"
            _upsert(session, variety_id, "seeder_roller_plate", None, value_text=seeder_text)
            upserted = True

        # Rows per bed
        rows_bed = _parse_int(row[_COL_ROWS_BED])
        if rows_bed is not None:
            _upsert(session, variety_id, "rows_per_bed", Decimal(rows_bed))
            upserted = True

        if upserted:
            summary.rows_upserted += 1

    summary.map_misses = map_misses
    logger.info(
        "Curtis L40: parsed=%d upserted=%d map_misses=%d",
        summary.rows_parsed, summary.rows_upserted, len(map_misses),
    )
    return summary
