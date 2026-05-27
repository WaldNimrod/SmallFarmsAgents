"""Idan 2018 diff importer (L49 vs L03/L04) — WP-C3 §9.

Compares L49 (2018 update) against L03/L04 (2017 baseline).
Upserts only new or changed rows with source='OP:Idan_2018'.
Writes diff report to _COMMUNICATION/team_10/SFA-S003-P002-WP-C3/L49_DIFF_REPORT.md.
"""

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from organic_market_agent.crop_book.importer.israeli._shared import (
    ImportSummary,
    resolve_crop_id,
)

logger = logging.getLogger(__name__)

SOURCE = "OP:Idan_2018"
TRUST = "OP"
CONFIDENCE = Decimal("0.55")

_REPO_ROOT = Path(__file__).resolve().parents[4]
_EXTERNAL_SOURCES_DIR = _REPO_ROOT / "data" / "external_sources"
_DIFF_REPORT_PATH = (
    _REPO_ROOT
    / "_COMMUNICATION"
    / "team_10"
    / "SFA-S003-P002-WP-C3"
    / "L49_DIFF_REPORT.md"
)

# Columns in sheet "תוכנית גידול" (0-based)
_SHEET = "תוכנית גידול"
_COL_CROP = 0
_COL_VARIETY = 1
_COL_AREA = 7       # שטח לזריעה (m²)
_COL_PLANTS_M2 = 9  # L49: מספר צמחים למ"ר
_COL_YIELD_KG = 12  # L49: כמות סה"כ(קילו)

# L03/L04 column positions differ slightly
_L03_COL_PLANTS_M2 = 11
_L03_COL_YIELD_KG = 14
_L03_COL_AREA = 7


def _parse_decimal(raw: Any) -> Decimal | None:
    if raw is None:
        return None
    try:
        return Decimal(str(raw)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _load_baseline(xlsx_path: Path, plants_col: int, yield_col: int, area_col: int) -> dict[str, dict[str, Decimal | None]]:
    """Load L03 or L04 → {crop_he: {field_name: value}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[_SHEET]
    rows_iter = ws.iter_rows(values_only=True)

    baseline: dict[str, dict[str, Decimal | None]] = {}
    header_found = False

    for row in rows_iter:
        if not row:
            continue
        if row[0] == "גידול":
            header_found = True
            continue
        if not header_found:
            continue
        crop_raw = row[0] if len(row) > 0 else None
        if not crop_raw:
            continue
        crop_he = str(crop_raw).strip()
        if not crop_he:
            continue

        plants_m2 = _parse_decimal(row[plants_col] if plants_col < len(row) else None)
        yield_kg = _parse_decimal(row[yield_col] if yield_col < len(row) else None)
        area = _parse_decimal(row[area_col] if area_col < len(row) else None)

        # Compute yield per m² if possible
        yield_per_m2: Decimal | None = None
        if yield_kg and area and area > 0:
            yield_per_m2 = (yield_kg / area).quantize(Decimal("0.01"))

        baseline.setdefault(crop_he, {})
        if plants_m2 is not None:
            baseline[crop_he]["plants_per_m2"] = plants_m2
        if yield_per_m2 is not None:
            baseline[crop_he]["yield_per_m2_kg"] = yield_per_m2

    return baseline


def _load_l49(xlsx_path: Path) -> list[dict]:
    """Load L49 → list of {crop_he, field_name, value_numeric}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[_SHEET]
    rows_iter = ws.iter_rows(values_only=True)

    results: list[dict] = []
    header_found = False

    for row in rows_iter:
        if not row:
            continue
        if row[0] == "גידול":
            header_found = True
            continue
        if not header_found:
            continue
        crop_raw = row[0] if len(row) > 0 else None
        if not crop_raw:
            continue
        crop_he = str(crop_raw).strip()
        if not crop_he:
            continue

        plants_m2 = _parse_decimal(row[_COL_PLANTS_M2] if _COL_PLANTS_M2 < len(row) else None)
        yield_kg = _parse_decimal(row[_COL_YIELD_KG] if _COL_YIELD_KG < len(row) else None)
        area = _parse_decimal(row[_COL_AREA] if _COL_AREA < len(row) else None)

        yield_per_m2: Decimal | None = None
        if yield_kg and area and area > 0:
            yield_per_m2 = (yield_kg / area).quantize(Decimal("0.01"))

        if plants_m2 is not None:
            results.append({"crop_he": crop_he, "field_name": "plants_per_m2", "value_numeric": plants_m2})
        if yield_per_m2 is not None:
            results.append({"crop_he": crop_he, "field_name": "yield_per_m2_kg", "value_numeric": yield_per_m2})

    return results


def _tolerance_match(v1: Decimal | None, v2: Decimal | None) -> bool:
    """Return True if two values are close enough to be considered identical."""
    if v1 is None and v2 is None:
        return True
    if v1 is None or v2 is None:
        return False
    diff = abs(v1 - v2)
    relative = diff / max(abs(v1), Decimal("0.0001"))
    return relative < Decimal("0.05")  # 5% tolerance


def _get_or_create_variety_id(session: Session, crop_id: int) -> int:
    from organic_market_agent.crop_book.models import CropVariety

    v = (
        session.query(CropVariety)
        .filter(CropVariety.crop_id == crop_id, CropVariety.name_en.is_(None))
        .order_by(CropVariety.id)
        .first()
    )
    if v is None:
        v = CropVariety(crop_id=crop_id, name_en=None, name_he=None)
        session.add(v)
        session.flush()
    return v.id


def import_all(
    session: Session,
    l49_path: Path | None = None,
    l03_path: Path | None = None,
    l04_path: Path | None = None,
) -> ImportSummary:
    l49 = l49_path or (_EXTERNAL_SOURCES_DIR / "israeli" / "L49_IDAN_market_gardening_tech.xlsx")
    l03 = l03_path or (_EXTERNAL_SOURCES_DIR / "israeli" / "L03_IDAN_winter_planning.xlsx")
    l04 = l04_path or (_EXTERNAL_SOURCES_DIR / "israeli" / "L04_IDAN_summer_planning.xlsx")

    summary = ImportSummary()

    # Build baseline from L03 + L04
    baseline: dict[str, dict[str, Decimal | None]] = {}
    for path, plants_col, yield_col in [
        (l03, _L03_COL_PLANTS_M2, _L03_COL_YIELD_KG),
        (l04, _L03_COL_PLANTS_M2, _L03_COL_YIELD_KG),
    ]:
        if path.exists():
            bl = _load_baseline(path, plants_col, yield_col, _L03_COL_AREA)
            for crop, fields in bl.items():
                baseline.setdefault(crop, {}).update(fields)

    # Load L49
    l49_rows = _load_l49(l49)
    summary.rows_parsed = len(l49_rows)

    changed_rows: list[dict] = []
    new_rows: list[dict] = []
    identical_rows: list[dict] = []

    from organic_market_agent.crop_book.models import CropVarietySourceValue

    for item in l49_rows:
        crop_he = item["crop_he"]
        field_name = item["field_name"]
        value_l49 = item["value_numeric"]

        baseline_val = baseline.get(crop_he, {}).get(field_name)

        if baseline_val is not None:
            if _tolerance_match(value_l49, baseline_val):
                identical_rows.append(item)
                continue
            else:
                changed_rows.append({**item, "baseline_value": baseline_val})
        else:
            new_rows.append(item)

        # Upsert this row
        crop_id, _ = resolve_crop_id(session, crop_he)
        if crop_id is None:
            summary.map_misses.append(crop_he)
            continue

        variety_id = _get_or_create_variety_id(session, crop_id)

        row = (
            session.query(CropVarietySourceValue)
            .filter_by(variety_id=variety_id, field_name=field_name, source=SOURCE)
            .one_or_none()
        )
        if row is None:
            row = CropVarietySourceValue(
                variety_id=variety_id,
                field_name=field_name,
                source=SOURCE,
            )
            session.add(row)
        row.value_numeric = value_l49
        row.value_text = str(value_l49)
        row.trust_tier = TRUST
        row.confidence_weight = CONFIDENCE
        row.is_outlier_rejected = False
        session.flush()
        summary.rows_upserted += 1

    # Write diff report
    _write_diff_report(changed_rows, new_rows, identical_rows, summary)

    logger.info(
        "Idan 2018 diff: parsed=%d changed=%d new=%d identical=%d upserted=%d",
        summary.rows_parsed, len(changed_rows), len(new_rows), len(identical_rows),
        summary.rows_upserted,
    )
    return summary


def _write_diff_report(
    changed: list[dict],
    new: list[dict],
    identical: list[dict],
    summary: ImportSummary,
) -> None:
    lines = [
        "# L49 Diff Report — Idan 2018 vs L03/L04 (2017)",
        "",
        f"**Generated:** 2026-05-27",
        f"**Source:** `OP:Idan_2018` (L49) vs `OP:Idan_2017` (L03/L04)",
        "",
        f"| Category | Count |",
        f"|----------|-------|",
        f"| Changed (L49 ≠ L03/L04) | {len(changed)} |",
        f"| New (not in L03/L04) | {len(new)} |",
        f"| Identical (skipped) | {len(identical)} |",
        f"| DB misses | {len(summary.map_misses)} |",
        f"| Total upserted | {summary.rows_upserted} |",
        "",
    ]

    if changed:
        lines += ["## Changed rows (L49 supersedes L03/L04)", ""]
        lines += ["| Crop | Field | L03/L04 value | L49 value |", "|------|-------|---------------|-----------|"]
        for row in changed:
            lines.append(f"| {row['crop_he']} | {row['field_name']} | {row.get('baseline_value')} | {row['value_numeric']} |")
        lines.append("")

    if new:
        lines += ["## New rows (not in L03/L04)", ""]
        lines += ["| Crop | Field | L49 value |", "|------|-------|-----------|"]
        for row in new:
            lines.append(f"| {row['crop_he']} | {row['field_name']} | {row['value_numeric']} |")
        lines.append("")

    if summary.map_misses:
        lines += ["## DB misses (crop not in DB)", ""]
        for miss in sorted(set(summary.map_misses)):
            lines.append(f"- `{miss}`")
        lines.append("")

    _DIFF_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    _DIFF_REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    logger.info("L49 diff report written to %s", _DIFF_REPORT_PATH)
