"""FRANCHI Italian seed catalog importer (L06 sheet 2 'גיליון2') — WP-C3 §8.

Source: L06_covers_and_tunnels.xlsx sheet 2 (27 rows × 5 cols; spec says 29 — discrepancy noted)
Target: crop_variety_source_values with source='OP:FRANCHI_catalog', field_name='variety_provider'
One row per crop (varieties concatenated, unique on variety_id/field/source per DB constraint).
"""

from __future__ import annotations

import logging
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from organic_market_agent.crop_book.importer.israeli._shared import (
    ImportSummary,
    resolve_crop_id,
)

logger = logging.getLogger(__name__)

SOURCE = "OP:FRANCHI_catalog"
TRUST = "OP"
CONFIDENCE = Decimal("0.55")
_SHEET = "גיליון2"

_REPO_ROOT = Path(__file__).resolve().parents[4]
_EXTERNAL_SOURCES_DIR = _REPO_ROOT / "data" / "external_sources"


def _get_or_create_variety_id(session: Session, crop_id: int) -> int:
    from organic_market_agent.crop_book.models import CropVariety

    v = (
        session.query(CropVariety)
        .filter(CropVariety.crop_id == crop_id, CropVariety.name_en.is_(None))
        .order_by(CropVariety.id)
        .first()
    )
    if v is None:
        v = CropVariety(crop_id=crop_id, name_en=None, name_he=None)
        session.add(v)
        session.flush()
    return v.id


def import_all(session: Session, xlsx_path: Path | None = None) -> ImportSummary:
    xlsx = xlsx_path or (_EXTERNAL_SOURCES_DIR / "israeli" / "L06_covers_and_tunnels.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[_SHEET]

    summary = ImportSummary()
    from organic_market_agent.crop_book.models import CropVarietySourceValue

    # Collect variety names per Hebrew crop name first (constraint is UNIQUE per variety_id/field/source)
    per_crop: dict[str, list[str]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or all(c is None for c in row):
            continue
        crop_he = str(row[0]).strip() if row[0] else ""
        variety_raw = str(row[1]).strip() if row[1] else ""
        code = str(row[2]).strip() if row[2] else ""

        if not crop_he or not variety_raw:
            continue

        summary.rows_parsed += 1
        value_text = f"{variety_raw} {code}".strip()
        per_crop[crop_he].append(value_text)

    # One upsert per crop (concatenated variety list)
    for crop_he, varieties in per_crop.items():
        crop_id, _ = resolve_crop_id(session, crop_he)
        if crop_id is None:
            summary.map_misses.append(crop_he)
            continue

        variety_id = _get_or_create_variety_id(session, crop_id)
        combined = " | ".join(varieties)

        existing = (
            session.query(CropVarietySourceValue)
            .filter_by(variety_id=variety_id, field_name="variety_provider", source=SOURCE)
            .one_or_none()
        )
        if existing is None:
            session.add(CropVarietySourceValue(
                variety_id=variety_id,
                field_name="variety_provider",
                source=SOURCE,
                value_text=combined,
                trust_tier=TRUST,
                confidence_weight=CONFIDENCE,
                is_outlier_rejected=False,
            ))
        else:
            existing.value_text = combined

        summary.rows_upserted += 1

    session.flush()
    logger.info(
        "FRANCHI L06: parsed=%d upserted=%d map_misses=%d",
        summary.rows_parsed, summary.rows_upserted, len(summary.map_misses),
    )
    return summary
