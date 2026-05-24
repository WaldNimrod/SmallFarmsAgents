"""JMF MasterClass XLSX importer (SFA-S003-P002-WP-B1 LOD400 §6).

Reads the JMF MasterClass workbook (5 sheets) and an optional pair of
standalone direct-seeding / nursery copies, producing:

  1. CropVarietySourceValue rows  (source='JMF', trust_tier='PR',
     confidence_weight=0.70) for 11 field names — fed to the WP-A
     enrichment engine via the standard upsert path.
  2. CropTaskTemplate rows (migration 044) for discrete growing tasks.

Public entrypoints:

  parse_crop_chart(xlsx_path)            -> list[dict]
  parse_associated_tasks(xlsx_path)      -> list[dict]
  parse_direct_seeding_chart(xlsx_path)  -> list[dict]
  parse_nursery_chart(xlsx_path)         -> list[dict]
  parse_cultivars(xlsx_path)             -> list[dict]
  import_jmf_masterclass(session, jmf_dir, *, dry_run=False) -> JmfImportSummary

All parsers return Python primitives only (no DB writes). The orchestrator
opens a transaction, calls upsert helpers, and commits.
"""

from __future__ import annotations

import glob
import logging
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class JmfImportSummary:
    crops_seen: int
    source_value_rows_upserted: int
    task_template_rows_upserted: int
    map_misses: list[str]    # JMF crop names with no JMF_CROP_MAP entry
    standalone_divergences: list[tuple[str, str, str, str]]
    # (sheet, crop_he, field_name, "<master>!=<standalone>")
    invalid_offsets: int = 0
    # F-S-002 (R1): count of CROP ASSOCIATED TASKS cells whose integer
    # value collided with DAYS_OFFSET_PRESENCE_ONLY (= -32768) and were
    # therefore skipped with ERROR. Expected: 0 in practice; this counter
    # exists to catch upstream data corruption early.


# ---------------------------------------------------------------------------
# Task column / timing maps (LOD400 §6.4)
# ---------------------------------------------------------------------------

_TASK_COLUMN_MAP: dict[str, str] = {
    "Stale Seed Bed":         "stale_seed_bed",
    "stale seed bed":         "stale_seed_bed",
    "Flame":                  "flame_weeder",
    "flame weeder":           "flame_weeder",
    "Flextine 1":             "flextine_harrow_1",
    "flextine harrow-1":      "flextine_harrow_1",
    "Flextine 2":             "flextine_harrow_2",
    "flextine harrow-2":      "flextine_harrow_2",
    "Biodisc":                "biodisc",
    "Hoe":                    "hoe",
    "Hand Weed":              "hand_weed",
    "handweed":               "hand_weed",
    "Boron Seaweed 1":        "boron_seaweed_1",
    "boron/seaweed -1":       "boron_seaweed_1",
    "Boron Seaweed 2":        "boron_seaweed_2",
    "boron/seaweed -2":       "boron_seaweed_2",
    "Straw Mulch":            "straw_mulch_topdress",
    "straw mulch+topdress":   "straw_mulch_topdress",
    "straw mulch":            "straw_mulch_topdress",
    "Head Pinch":             "head_pinch_chop",
    "head pinch/chop":        "head_pinch_chop",
    "Mow and Tarp":           "mow_and_tarp",
    "mow and tarp":           "mow_and_tarp",
    "At Seeding":             "at_seeding_transplanting",
    "at seeding/transplanting": "at_seeding_transplanting",
    "Row Cover":              "net_row_cover",
    "net/row cover":          "net_row_cover",
    "net row cover":          "net_row_cover",
}

_TASK_TIMING_MAP: dict[str, str] = {
    "stale_seed_bed":  "field_prep",
    "flame_weeder":    "field_prep",
    "at_seeding_transplanting": "seeding",
    # All others default to "seeding".
}


def _match_task_column(header: str) -> Optional[str]:
    """Case-insensitive substring match for task column headers."""
    if not header:
        return None
    h_lower = header.lower().strip()
    # Exact match first (case-insensitive)
    for key, value in _TASK_COLUMN_MAP.items():
        if key.lower() == h_lower:
            return value
    # Substring match
    for key, value in _TASK_COLUMN_MAP.items():
        if key.lower() in h_lower or h_lower in key.lower():
            return value
    return None


# ---------------------------------------------------------------------------
# Unit conversion helpers (LOD400 §7)
# ---------------------------------------------------------------------------

_YIELD_CONVERSIONS: dict[str, Decimal] = {
    "lbs/100ft": Decimal("0.453592") / Decimal("30.48"),
    "kg/100m": Decimal("0.01"),
    "kg/100ft": Decimal("1") / Decimal("30.48"),
    "lbs/100m": Decimal("0.453592") / Decimal("100"),
    # Aliases
    "lb/100ft": Decimal("0.453592") / Decimal("30.48"),
    "lb/100'": Decimal("0.453592") / Decimal("30.48"),
    "lbs/100'": Decimal("0.453592") / Decimal("30.48"),
    "lbs/100ft bed": Decimal("0.453592") / Decimal("30.48"),
}


def _yield_to_per_meter(
    value_raw: object,
    unit_raw: str = "lbs/100ft",
) -> Optional[Decimal]:
    """Convert yield to kg per 1 meter of bed. Returns None on NULL/invalid."""
    if value_raw is None or str(value_raw).strip() == "":
        return None
    try:
        value = Decimal(str(value_raw).replace(",", "."))
    except (InvalidOperation, ValueError):
        logger.warning("yield_to_per_meter: cannot parse value %r", value_raw)
        return None

    unit_key = str(unit_raw).strip().lower() if unit_raw else "lbs/100ft"
    multiplier = None
    # Exact match
    for k, m in _YIELD_CONVERSIONS.items():
        if k.lower() == unit_key:
            multiplier = m
            break
    # Substring match
    if multiplier is None:
        for k, m in _YIELD_CONVERSIONS.items():
            if k.lower() in unit_key or unit_key in k.lower():
                multiplier = m
                break
    if multiplier is None:
        logger.warning("yield_to_per_meter: unknown unit %r — skipping", unit_raw)
        return None

    result = value * multiplier
    return result.quantize(Decimal("0.0001"))


def _inches_to_cm(value_raw: object) -> Optional[Decimal]:
    """Convert inches to cm. Returns None on NULL/blank/invalid."""
    if value_raw is None:
        return None
    s = str(value_raw).strip()
    if not s:
        return None
    # Strip common inch notation: 2", 2'', 2,5", etc.
    s = s.replace('"', '').replace("'", '').replace(',', '.').strip()
    # Sometimes like "1.5" or "1,5"
    try:
        value = Decimal(s)
    except (InvalidOperation, ValueError):
        logger.warning("inches_to_cm: cannot parse %r", value_raw)
        return None
    result = value * Decimal("2.54")
    return result.quantize(Decimal("0.01"))


def _safe_int(value: object) -> Optional[int]:
    """Safely convert a cell value to int. Returns None on failure."""
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return None


def _safe_decimal(value: object) -> Optional[Decimal]:
    """Safely convert a cell value to Decimal. Returns None on failure."""
    if value is None:
        return None
    s = str(value).strip().replace(',', '.').lstrip('$').lstrip('€').lstrip('£')
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------------
# Column header finder (case-insensitive substring)
# ---------------------------------------------------------------------------

def _find_col(headers: list, fragment: str) -> Optional[int]:
    """Return the column index whose header contains `fragment` (case-insensitive)."""
    frag_lower = fragment.lower()
    for i, h in enumerate(headers):
        if h and frag_lower in str(h).lower():
            return i
    return None


# ---------------------------------------------------------------------------
# 5 Sheet parsers
# ---------------------------------------------------------------------------

def _load_sheet(xlsx_path: Path, sheet_hints: list[str]):
    """Load workbook and return the best-matching sheet by name hints."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    for hint in sheet_hints:
        h_lower = hint.lower().strip()
        for name in wb.sheetnames:
            if h_lower in name.lower():
                return wb[name]
    # fallback: first sheet
    return wb[wb.sheetnames[0]]


def parse_crop_chart(xlsx_path: Path) -> list[dict]:
    """Parse CROP CHART sheet. Returns list[dict] with crop-scoped field observations.

    LOD400 §6.3. Fields: crop_jmf_en, days_to_maturity, harvest_window_max_days,
    avg_yield_per_bed_m, documented_price, documented_price_unit.
    """
    xlsx_path = Path(xlsx_path)
    ws = _load_sheet(xlsx_path, ["CROP CHART"])

    rows_data = list(ws.iter_rows(values_only=True))
    headers = None
    results = []

    for row in rows_data:
        # Find header row: must contain 'Crop' or 'crop' in first column or near start
        if headers is None:
            row_strs = [str(c).strip().lower() if c else "" for c in row]
            if "crop" in row_strs[0:3]:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            continue

        if not row[0]:
            continue
        crop_name = str(row[0]).strip()
        if not crop_name:
            continue

        col_crop = _find_col(headers, "crop") if headers else 0
        col_dtm = _find_col(headers, "DTM")
        if col_dtm is None:
            col_dtm = _find_col(headers, "Days to Maturity")
        col_hw = _find_col(headers, "Harvest window")
        if col_hw is None:
            col_hw = _find_col(headers, "harvest window")
        col_yield = _find_col(headers, "Yield")
        if col_yield is None:
            col_yield = _find_col(headers, "yield")
        col_price = _find_col(headers, "Price")
        if col_price is None:
            col_price = _find_col(headers, "price")
        col_unit = _find_col(headers, "Unit")
        if col_unit is None:
            col_unit = _find_col(headers, "unit")

        entry: dict = {"crop_jmf_en": crop_name}

        # DTM
        if col_dtm is not None and col_dtm < len(row):
            v = _safe_int(row[col_dtm])
            if v is not None:
                entry["days_to_maturity"] = v

        # Harvest window (max)
        if col_hw is not None and col_hw < len(row):
            v = _safe_int(row[col_hw])
            if v is not None:
                entry["harvest_window_max_days"] = v

        # Yield → per-meter conversion
        if col_yield is not None and col_yield < len(row):
            # Try to determine unit from header or a nearby unit header
            yield_raw = row[col_yield]
            # Default unit assumption: lbs/100ft (JMF default)
            unit_str = "lbs/100ft"
            if col_unit is not None and col_unit < len(row) and row[col_unit]:
                u = str(row[col_unit]).strip().lower()
                # If unit col says Lbs, kg etc — map to yield unit
                if "kg" in u:
                    unit_str = "kg/100ft"
                elif "lb" in u or "lbs" in u:
                    unit_str = "lbs/100ft"
            converted = _yield_to_per_meter(yield_raw, unit_str)
            if converted is not None:
                entry["avg_yield_per_bed_m"] = converted

        # Price
        if col_price is not None and col_price < len(row):
            v = _safe_decimal(row[col_price])
            if v is not None:
                entry["documented_price"] = v
                # Unit for price
                if col_unit is not None and col_unit < len(row) and row[col_unit]:
                    entry["documented_price_unit"] = str(row[col_unit]).strip()

        results.append(entry)

    return results


def parse_associated_tasks(xlsx_path: Path) -> list[dict]:
    """Parse CROP ASSOCIATED TASKS sheet. Returns task-template dicts.

    LOD400 §6.4. Each non-blank cell → one dict with crop_jmf_en, task_type,
    timing_anchor, days_offset (int; DAYS_OFFSET_PRESENCE_ONLY for X cells).
    """
    from organic_market_agent.crop_book.crop_task_templates import DAYS_OFFSET_PRESENCE_ONLY

    xlsx_path = Path(xlsx_path)
    ws = _load_sheet(xlsx_path, ["CROP ASSOCIATED TASKS"])

    rows_data = list(ws.iter_rows(values_only=True))
    headers = None
    task_col_map: dict[int, str] = {}  # col_index → task_type
    crop_col_idx: int = 1  # default: 2nd column
    results = []
    invalid_offsets = 0

    for row in rows_data:
        # Find header row — must contain 'crop' in positions 0-2 AND at least one known task column
        if headers is None:
            row_strs = [str(c).strip().lower() if c else "" for c in row]
            has_crop = any("crop" in s for s in row_strs[:3])
            has_task = any(
                any(frag in s for frag in ["seed bed", "flame weeder", "flextine", "biodisc",
                                            "hand weed", "handweed", "boron", "straw mulch",
                                            "head pinch", "mow and tarp", "at seeding",
                                            "row cover", "net/row"])
                for s in row_strs
            )
            if has_crop and has_task:
                headers = [str(c).strip() if c else "" for c in row]
                # Find crop column
                for i, h in enumerate(headers):
                    if h and "crop" in h.lower():
                        crop_col_idx = i
                        break
                # Map other columns to task types
                for i, h in enumerate(headers):
                    if i == crop_col_idx:
                        continue
                    if not h:
                        continue
                    tt = _match_task_column(h)
                    if tt:
                        task_col_map[i] = tt
                    elif h.strip():
                        logger.debug("parse_associated_tasks: unmapped column header %r", h)
                continue

        if not row or crop_col_idx >= len(row):
            continue
        crop_raw = row[crop_col_idx]
        if not crop_raw or not str(crop_raw).strip():
            continue
        crop_name = str(crop_raw).strip()

        # Each non-empty task cell
        for col_idx, task_type in task_col_map.items():
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            if cell_val is None or str(cell_val).strip() == "":
                continue

            cell_str = str(cell_val).strip()

            # X / presence-only
            if cell_str.upper() == "X":
                days_off = DAYS_OFFSET_PRESENCE_ONLY
                notes = None
            else:
                # Try integer parse
                try:
                    int_val = int(cell_str)
                    if int_val == DAYS_OFFSET_PRESENCE_ONLY:
                        logger.error(
                            "parse_associated_tasks: cell value equals sentinel "
                            "(-32768) for crop=%r task_type=%r — skipping",
                            crop_name, task_type,
                        )
                        invalid_offsets += 1
                        continue
                    days_off = int_val
                    notes = None
                except ValueError:
                    # Non-numeric text → notes + sentinel
                    logger.warning(
                        "parse_associated_tasks: non-numeric cell %r for crop=%r task_type=%r "
                        "— storing in notes, using presence-only sentinel",
                        cell_str, crop_name, task_type,
                    )
                    days_off = DAYS_OFFSET_PRESENCE_ONLY
                    notes = cell_str

            timing_anchor = _TASK_TIMING_MAP.get(task_type, "seeding")

            results.append({
                "crop_jmf_en": crop_name,
                "task_type": task_type,
                "timing_anchor": timing_anchor,
                "days_offset": days_off,
                "method": None,
                "input_material": None,
                "notes": notes,
                "_invalid_offsets": invalid_offsets,
            })

    # Embed invalid_offsets count in the last entry or via a side-channel
    # (callers read it from the summary; we pass it through records)
    return results


def _get_invalid_offsets_from_tasks(task_rows: list[dict]) -> int:
    """Extract invalid_offsets count from task rows."""
    if not task_rows:
        return 0
    return task_rows[-1].get("_invalid_offsets", 0)


def parse_direct_seeding_chart(xlsx_path: Path) -> list[dict]:
    """Parse DIRECT SEEDING CHART sheet. Returns list[dict].

    LOD400 §6.5. Fields: crop_jmf_en, in_row_spacing_cm, rows_per_bed,
    direct_seed_density_g, seeder.
    """
    xlsx_path = Path(xlsx_path)
    ws = _load_sheet(xlsx_path, ["DIRECT SEEDING CHART", "DIRECT SEEDING"])

    rows_data = list(ws.iter_rows(values_only=True))
    headers = None
    results = []

    for row in rows_data:
        if headers is None:
            row_strs = [str(c).strip().lower() if c else "" for c in row[:5]]
            if "crop" in row_strs:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            continue

        if not row[0]:
            continue
        crop_name = str(row[0]).strip()
        if not crop_name:
            continue

        col_rows = _find_col(headers, "Rows per Bed")
        if col_rows is None:
            col_rows = _find_col(headers, "ROWS PER BED")
        if col_rows is None:
            col_rows = _find_col(headers, "# Row")
        if col_rows is None:
            col_rows = _find_col(headers, "row")

        col_spacing = _find_col(headers, "In-Row Spacing")
        if col_spacing is None:
            col_spacing = _find_col(headers, "SPACING")
        if col_spacing is None:
            col_spacing = _find_col(headers, "spacing")

        col_density = _find_col(headers, "Seed Density")
        if col_density is None:
            col_density = _find_col(headers, "DENSITY")
        if col_density is None:
            col_density = _find_col(headers, "density")

        col_seeder = _find_col(headers, "Seeder")
        if col_seeder is None:
            col_seeder = _find_col(headers, "SEEDER")

        entry: dict = {"crop_jmf_en": crop_name}

        if col_spacing is not None and col_spacing < len(row):
            cm = _inches_to_cm(row[col_spacing])
            if cm is not None:
                entry["in_row_spacing_cm"] = cm

        if col_rows is not None and col_rows < len(row):
            v = _safe_int(row[col_rows])
            if v is not None:
                entry["rows_per_bed"] = v

        if col_density is not None and col_density < len(row):
            raw = row[col_density]
            if raw is not None:
                # Density is often "90 g" — strip unit
                s = str(raw).strip().replace('g', '').replace(',', '.').strip()
                try:
                    entry["direct_seed_density_g"] = Decimal(s)
                except (InvalidOperation, ValueError):
                    pass

        if col_seeder is not None and col_seeder < len(row):
            v = row[col_seeder]
            if v and str(v).strip():
                entry["seeder"] = str(v).strip()

        results.append(entry)

    return results


def parse_nursery_chart(xlsx_path: Path) -> list[dict]:
    """Parse NURSERY & TRANSPLANT CHART sheet. Returns list[dict].

    LOD400 §6.6. Fields: crop_jmf_en, days_in_nursery_cell, nursery_tray_type,
    in_row_spacing_cm, rows_per_bed.
    """
    xlsx_path = Path(xlsx_path)
    ws = _load_sheet(xlsx_path, ["NURSERY", "TRANSPLANT CHART"])

    rows_data = list(ws.iter_rows(values_only=True))
    headers = None
    results = []

    for row in rows_data:
        if headers is None:
            row_strs = [str(c).strip().lower() if c else "" for c in row[:5]]
            if "crop" in row_strs:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            continue

        if not row[0]:
            continue
        crop_name = str(row[0]).strip()
        if not crop_name:
            continue

        col_days = _find_col(headers, "Days in Cell")
        if col_days is None:
            col_days = _find_col(headers, "Days in cell")
        if col_days is None:
            col_days = _find_col(headers, "days in cell")

        col_tray = _find_col(headers, "Tray")
        if col_tray is None:
            col_tray = _find_col(headers, "tray")
        if col_tray is None:
            col_tray = _find_col(headers, "Tray number")

        col_spacing = _find_col(headers, "Spacing on the row")
        if col_spacing is None:
            col_spacing = _find_col(headers, "In-Row Spacing")
        if col_spacing is None:
            col_spacing = _find_col(headers, "Spacing")

        col_rows = _find_col(headers, "# Row")
        if col_rows is None:
            col_rows = _find_col(headers, "Rows per Bed")
        if col_rows is None:
            col_rows = _find_col(headers, "# Row")

        entry: dict = {"crop_jmf_en": crop_name}

        # Days in cell — stored as midpoint if range
        if col_days is not None and col_days < len(row):
            raw = row[col_days]
            if raw is not None:
                s = str(raw).strip()
                if '-' in s:
                    try:
                        parts = [int(x.strip()) for x in s.split('-')]
                        mid = round(sum(parts) / len(parts))
                        entry["days_in_nursery_cell"] = mid
                        entry.setdefault("_nursery_note", f"range:{parts[0]}-{parts[-1]}")
                    except ValueError:
                        pass
                else:
                    v = _safe_int(raw)
                    if v is not None:
                        entry["days_in_nursery_cell"] = v

        if col_tray is not None and col_tray < len(row):
            v = row[col_tray]
            if v and str(v).strip():
                entry["nursery_tray_type"] = str(v).strip()

        if col_spacing is not None and col_spacing < len(row):
            cm = _inches_to_cm(row[col_spacing])
            if cm is not None:
                entry["in_row_spacing_cm"] = cm

        if col_rows is not None and col_rows < len(row):
            v = _safe_int(row[col_rows])
            if v is not None:
                entry["rows_per_bed"] = v

        results.append(entry)

    return results


def parse_cultivars(xlsx_path: Path) -> list[dict]:
    """Parse CULTIVARS sheet. Returns list[dict].

    LOD400 §6.7. Fields: crop_jmf_en, variety_name_en, cultivar_provider,
    days_to_maturity, cultivar_description.
    """
    xlsx_path = Path(xlsx_path)
    ws = _load_sheet(xlsx_path, ["CULTIVARS"])

    rows_data = list(ws.iter_rows(values_only=True))
    headers = None
    results = []
    current_crop = None  # carry-forward for merged cells

    for row in rows_data:
        if headers is None:
            row_strs = [str(c).strip().lower() if c else "" for c in row[:5]]
            if "species" in row_strs or "crop" in row_strs or "cultivar" in row_strs:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            continue

        # Resolve current crop (carry-forward for merged-cell layout)
        crop_raw = row[0]
        if crop_raw and str(crop_raw).strip():
            current_crop = str(crop_raw).strip()
        if not current_crop:
            continue

        col_cultivar = _find_col(headers, "Cultivar")
        if col_cultivar is None:
            col_cultivar = 1  # default

        col_provider = _find_col(headers, "Provider")
        if col_provider is None:
            col_provider = _find_col(headers, "Supplier")
        if col_provider is None:
            col_provider = 2

        col_dtm = _find_col(headers, "DTM")

        col_desc = _find_col(headers, "Description")
        if col_desc is None:
            col_desc = _find_col(headers, "Information")
        if col_desc is None:
            col_desc = 5

        col_comments = _find_col(headers, "Comments")
        if col_comments is None:
            col_comments = 6

        # Skip rows where cultivar column is empty
        cultivar_val = row[col_cultivar] if col_cultivar < len(row) else None
        if not cultivar_val or not str(cultivar_val).strip():
            continue

        entry: dict = {
            "crop_jmf_en": current_crop,
            "variety_name_en": str(cultivar_val).strip(),
        }

        if col_provider < len(row) and row[col_provider]:
            entry["cultivar_provider"] = str(row[col_provider]).strip()

        if col_dtm is not None and col_dtm < len(row):
            v = _safe_int(row[col_dtm])
            if v is not None:
                entry["days_to_maturity"] = v

        # Description + Comments joined
        desc_parts = []
        if col_desc < len(row) and row[col_desc]:
            desc_parts.append(str(row[col_desc]).strip())
        if col_comments is not None and col_comments < len(row) and row[col_comments]:
            desc_parts.append(str(row[col_comments]).strip())
        if desc_parts:
            entry["cultivar_description"] = " / ".join(desc_parts)

        results.append(entry)

    return results


# ---------------------------------------------------------------------------
# Variety resolution helper
# ---------------------------------------------------------------------------

def _default_variety_id(session: Session, crop_id: int) -> int:
    """Get or create the baseline (default, name_en=None) variety for a crop."""
    from organic_market_agent.crop_book.models import CropVariety
    v = (session.query(CropVariety)
         .filter(CropVariety.crop_id == crop_id, CropVariety.name_en.is_(None))
         .one_or_none())
    if v is None:
        v = CropVariety(crop_id=crop_id, name_en=None, name_he=None)
        session.add(v)
        session.flush()
    return v.id


# ---------------------------------------------------------------------------
# Upsert helpers (LOD400 §6.10, §6.6)
# ---------------------------------------------------------------------------

def _upsert_source_value(
    session: Session,
    variety_id: int,
    field_name: str,
    value_numeric: Optional[Decimal] = None,
    value_text: Optional[str] = None,
    unit: Optional[str] = None,
    note: Optional[str] = None,
):
    """Upsert on (variety_id, field_name, source='JMF').

    trust_tier='PR', confidence_weight=0.70, is_outlier_rejected=False
    are hardcoded here — they are the contract this importer exists to provide.
    """
    from organic_market_agent.crop_book.models import CropVarietySourceValue

    SOURCE = "JMF"
    row = (session.query(CropVarietySourceValue)
           .filter_by(variety_id=variety_id, field_name=field_name, source=SOURCE)
           .one_or_none())
    if row is None:
        row = CropVarietySourceValue(
            variety_id=variety_id, field_name=field_name, source=SOURCE,
        )
        session.add(row)
    row.value_numeric = value_numeric
    row.value_text = value_text
    row.unit = unit
    row.note = note
    row.trust_tier = "PR"
    row.confidence_weight = Decimal("0.70")
    row.is_outlier_rejected = False
    session.flush()
    return row


def _upsert_task_template(
    session: Session,
    crop_id: int,
    task_row: dict,
) -> bool:
    """Upsert on (crop_id, source='JMF', task_type, days_offset). Returns True if new."""
    from organic_market_agent.crop_book.crop_task_templates import CropTaskTemplate

    SOURCE = "JMF"
    days_off = task_row["days_offset"]
    task_type = task_row["task_type"]

    existing = (session.query(CropTaskTemplate)
                .filter_by(crop_id=crop_id, source=SOURCE,
                            task_type=task_type, days_offset=days_off)
                .one_or_none())
    if existing is None:
        obj = CropTaskTemplate(
            crop_id=crop_id,
            source=SOURCE,
            trust_tier="PR",
            task_type=task_type,
            timing_anchor=task_row.get("timing_anchor"),
            days_offset=days_off,
            method=task_row.get("method"),
            input_material=task_row.get("input_material"),
            notes=task_row.get("notes"),
        )
        session.add(obj)
        session.flush()
        return True
    else:
        # Update mutable fields
        existing.timing_anchor = task_row.get("timing_anchor")
        existing.notes = task_row.get("notes")
        session.flush()
        return False


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def import_jmf_masterclass(
    session: Session,
    jmf_dir: Path,
    *,
    dry_run: bool = False,
) -> JmfImportSummary:
    """End-to-end orchestrator.

    Steps:
      1. Resolve the master workbook and standalone files.
      2. Call the 5 parsers (master) + 2 standalone parsers.
      3. Cross-check standalone vs master; master wins.
      4. Resolve crop_id via JMF_CROP_MAP. Miss → map_misses.
      5. Upsert source values via _upsert_source_value.
      6. Upsert task templates via _upsert_task_template.
      7. If dry_run: rollback; else commit.
      8. Return JmfImportSummary.
    """
    from organic_market_agent.crop_book.constants import JMF_CROP_MAP
    from organic_market_agent.crop_book.models import Crop, CropVariety

    jmf_dir = Path(jmf_dir)

    # Step 1: resolve master workbook
    master_candidates = sorted(glob.glob(str(jmf_dir / "CROPPLANNINGTOOLMASTERCLASS-*.XLSX")))
    if not master_candidates:
        # Case-insensitive fallback
        master_candidates = sorted(glob.glob(str(jmf_dir / "*.XLSX"))) + \
                            sorted(glob.glob(str(jmf_dir / "*.xlsx")))
    if not master_candidates:
        logger.warning("import_jmf_masterclass: no master workbook found in %s", jmf_dir)
        return JmfImportSummary(
            crops_seen=0, source_value_rows_upserted=0,
            task_template_rows_upserted=0, map_misses=[],
            standalone_divergences=[], invalid_offsets=0,
        )
    master_path = Path(master_candidates[0])
    logger.info("import_jmf_masterclass: using master %s", master_path.name)

    # Step 2: parse all 5 sheets from master
    crop_chart_rows = parse_crop_chart(master_path)
    task_rows = parse_associated_tasks(master_path)
    ds_rows_master = parse_direct_seeding_chart(master_path)
    nursery_rows_master = parse_nursery_chart(master_path)
    cultivar_rows = parse_cultivars(master_path)

    # Extract invalid_offsets from task rows
    invalid_offsets_count = _get_invalid_offsets_from_tasks(task_rows)
    # Clean up the internal counter key
    for tr in task_rows:
        tr.pop("_invalid_offsets", None)

    # Step 2b: standalone files
    standalone_dir = jmf_dir.parent / "תבלאות נתונים"
    ds_rows_standalone: list[dict] = []
    nursery_rows_standalone: list[dict] = []

    if standalone_dir.exists():
        ds_files = sorted(glob.glob(str(standalone_dir / "DIRECTSEEDINGCHART-*.XLSX")))
        if ds_files:
            ds_rows_standalone = parse_direct_seeding_chart(Path(ds_files[0]))
        nursery_files = sorted(glob.glob(str(standalone_dir / "NURSERYTRANSPLANTCHART-*.XLSX")))
        if nursery_files:
            nursery_rows_standalone = parse_nursery_chart(Path(nursery_files[0]))

    # Step 3: cross-check standalone vs master
    standalone_divergences: list[tuple[str, str, str, str]] = []
    _master_ds_by_crop = {r["crop_jmf_en"]: r for r in ds_rows_master}
    _master_ns_by_crop = {r["crop_jmf_en"]: r for r in nursery_rows_master}

    for r in ds_rows_standalone:
        crop = r["crop_jmf_en"]
        master_r = _master_ds_by_crop.get(crop)
        if master_r:
            for fld in ("in_row_spacing_cm", "rows_per_bed", "direct_seed_density_g"):
                mv = master_r.get(fld)
                sv = r.get(fld)
                if mv is not None and sv is not None and mv != sv:
                    crop_he = JMF_CROP_MAP.get(crop, crop)
                    standalone_divergences.append((
                        "DIRECT_SEEDING", crop_he, fld,
                        f"{mv}!={sv}",
                    ))

    for r in nursery_rows_standalone:
        crop = r["crop_jmf_en"]
        master_r = _master_ns_by_crop.get(crop)
        if master_r:
            for fld in ("in_row_spacing_cm", "rows_per_bed", "days_in_nursery_cell"):
                mv = master_r.get(fld)
                sv = r.get(fld)
                if mv is not None and sv is not None and mv != sv:
                    crop_he = JMF_CROP_MAP.get(crop, crop)
                    standalone_divergences.append((
                        "NURSERY", crop_he, fld,
                        f"{mv}!={sv}",
                    ))

    if standalone_divergences:
        logger.warning("Standalone divergences (%d) — master wins", len(standalone_divergences))

    # Step 4-6: resolve crops and upsert
    map_misses: list[str] = []
    seen_crops: set[str] = set()
    sv_upserted = 0
    tt_upserted = 0

    # Merge direct seeding and nursery into crop chart (master wins)
    ds_by_crop = {r["crop_jmf_en"]: r for r in ds_rows_master}
    ns_by_crop = {r["crop_jmf_en"]: r for r in nursery_rows_master}
    cult_by_crop: dict[str, list[dict]] = {}
    for r in cultivar_rows:
        cult_by_crop.setdefault(r["crop_jmf_en"], []).append(r)

    # Process crop-scoped rows
    all_crop_names: set[str] = set()
    for r in crop_chart_rows:
        all_crop_names.add(r["crop_jmf_en"])
    for r in task_rows:
        all_crop_names.add(r["crop_jmf_en"])
    for r in ds_rows_master:
        all_crop_names.add(r["crop_jmf_en"])
    for r in nursery_rows_master:
        all_crop_names.add(r["crop_jmf_en"])

    for crop_jmf_en in sorted(all_crop_names):
        # Resolve Hebrew name
        name_he = JMF_CROP_MAP.get(crop_jmf_en)
        if name_he is None:
            if crop_jmf_en not in map_misses:
                logger.warning("JMF map miss: %r not in JMF_CROP_MAP — skipping", crop_jmf_en)
                map_misses.append(crop_jmf_en)
            continue

        seen_crops.add(crop_jmf_en)

        # Get or create crop
        crop_obj = session.query(Crop).filter_by(name_he=name_he).one_or_none()
        if crop_obj is None:
            # For JMF-only crops not in the DB: create with minimal fields.
            # `family_id` is NOT NULL in the schema — use the first available family
            # as a placeholder, or skip if none exists.
            from organic_market_agent.crop_book.models import CropFamily
            family = session.query(CropFamily).first()
            if family is None:
                logger.warning(
                    "Cannot create new crop %r — no family exists in DB. Skipping.", name_he
                )
                if name_he not in map_misses:
                    map_misses.append(f"[no_family]:{name_he}")
                continue
            crop_obj = Crop(name_he=name_he, category="vegetables", family_id=family.id)
            session.add(crop_obj)
            session.flush()

        crop_id = crop_obj.id
        variety_id = _default_variety_id(session, crop_id)

        # Source values from CROP CHART
        cc_row = next((r for r in crop_chart_rows if r["crop_jmf_en"] == crop_jmf_en), None)
        if cc_row:
            for fld_name in ("days_to_maturity", "harvest_window_max_days",
                              "avg_yield_per_bed_m", "documented_price"):
                val = cc_row.get(fld_name)
                if val is not None:
                    unit = None
                    if fld_name == "days_to_maturity":
                        unit = "days"
                        val = Decimal(str(val))
                    elif fld_name == "harvest_window_max_days":
                        unit = "days"
                        val = Decimal(str(val))
                    elif fld_name == "avg_yield_per_bed_m":
                        unit = "kg/m"
                        # val is already Decimal
                    elif fld_name == "documented_price":
                        unit = cc_row.get("documented_price_unit", "unit")
                    _upsert_source_value(session, variety_id, fld_name,
                                         value_numeric=val, unit=unit)
                    sv_upserted += 1

        # Source values from DIRECT SEEDING
        ds_row = ds_by_crop.get(crop_jmf_en)
        if ds_row:
            for fld_name in ("in_row_spacing_cm", "rows_per_bed"):
                val = ds_row.get(fld_name)
                if val is not None:
                    _upsert_source_value(session, variety_id, fld_name,
                                         value_numeric=Decimal(str(val)),
                                         unit="cm" if "cm" in fld_name else None)
                    sv_upserted += 1
            density = ds_row.get("direct_seed_density_g")
            if density is not None:
                _upsert_source_value(session, variety_id, "direct_seed_density_g",
                                     value_numeric=density, unit="g/bed")
                sv_upserted += 1
            seeder = ds_row.get("seeder")
            if seeder:
                _upsert_source_value(session, variety_id, "seeder",
                                     value_text=seeder)
                sv_upserted += 1

        # Source values from NURSERY
        ns_row = ns_by_crop.get(crop_jmf_en)
        if ns_row:
            days_cell = ns_row.get("days_in_nursery_cell")
            if days_cell is not None:
                note = ns_row.get("_nursery_note")
                _upsert_source_value(session, variety_id, "days_in_nursery_cell",
                                     value_numeric=Decimal(str(days_cell)),
                                     unit="days", note=note)
                sv_upserted += 1
            tray = ns_row.get("nursery_tray_type")
            if tray:
                _upsert_source_value(session, variety_id, "nursery_tray_type",
                                     value_text=tray)
                sv_upserted += 1
            # in_row_spacing_cm from nursery (may overlap with DS)
            spc = ns_row.get("in_row_spacing_cm")
            if spc is not None:
                _upsert_source_value(session, variety_id, "in_row_spacing_cm",
                                     value_numeric=Decimal(str(spc)), unit="cm")
                # Not counted as additional upsert (possible overlap with DS)

            rows_per = ns_row.get("rows_per_bed")
            if rows_per is not None:
                _upsert_source_value(session, variety_id, "rows_per_bed",
                                     value_numeric=Decimal(str(rows_per)))

        # Task templates
        for tr in task_rows:
            if tr["crop_jmf_en"] != crop_jmf_en:
                continue
            is_new = _upsert_task_template(session, crop_id, tr)
            if is_new:
                tt_upserted += 1

    # Process CULTIVARS (variety-scoped)
    for crop_jmf_en, cult_list in cult_by_crop.items():
        name_he = JMF_CROP_MAP.get(crop_jmf_en)
        if name_he is None:
            if crop_jmf_en not in map_misses:
                map_misses.append(crop_jmf_en)
            continue

        crop_obj = session.query(Crop).filter_by(name_he=name_he).one_or_none()
        if crop_obj is None:
            from organic_market_agent.crop_book.models import CropFamily
            family = session.query(CropFamily).first()
            if family is None:
                continue
            crop_obj = Crop(name_he=name_he, category="vegetables", family_id=family.id)
            session.add(crop_obj)
            session.flush()
        crop_id = crop_obj.id

        for cult in cult_list:
            vname = cult.get("variety_name_en")
            if not vname:
                continue
            # Get or create variety by (crop_id, name_en)
            variety = (session.query(CropVariety)
                        .filter_by(crop_id=crop_id, name_en=vname)
                        .one_or_none())
            if variety is None:
                variety = CropVariety(crop_id=crop_id, name_en=vname)
                session.add(variety)
                session.flush()
            variety_id = variety.id

            if "days_to_maturity" in cult:
                _upsert_source_value(session, variety_id, "days_to_maturity",
                                     value_numeric=Decimal(str(cult["days_to_maturity"])),
                                     unit="days")
                sv_upserted += 1
            if "cultivar_provider" in cult:
                _upsert_source_value(session, variety_id, "cultivar_provider",
                                     value_text=cult["cultivar_provider"])
                sv_upserted += 1
            if "cultivar_description" in cult:
                _upsert_source_value(session, variety_id, "cultivar_description",
                                     value_text=cult["cultivar_description"])
                sv_upserted += 1

    # Step 7: rollback or leave for caller
    if dry_run:
        session.rollback()

    return JmfImportSummary(
        crops_seen=len(seen_crops),
        source_value_rows_upserted=sv_upserted,
        task_template_rows_upserted=tt_upserted,
        map_misses=map_misses,
        standalone_divergences=standalone_divergences,
        invalid_offsets=invalid_offsets_count,
    )
